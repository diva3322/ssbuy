import json
import time
from openpyxl import load_workbook # 用於讀取 .xlsx 檔案
import requests # 用於發送 HTTP 請求 (Google Custom Search API)
import os # 用於檢查檔案是否存在
import re # 用於正則表達式清理檔案名
import sys # 用於標準輸出設定
import io # 用於標準輸出設定

# 確保標準輸出能正確顯示中文
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# --- 請設定您的 Google Custom Search API 相關資訊 ---
API_KEY = "AIzaSyAapqNr8NGy0Br0D9K-qOPFXsvzYH9feDY" # 您的 Google API Key
SEARCH_ENGINE_ID = "5534bc0e180af4cf5" # 您的 Google Custom Search Engine ID
# --- 設定結束 ---

def clean_filename(name):
    """
    清理遊戲名稱，移除不適合作為檔案名的字元。
    """
    return re.sub(r'[\\/*?:"<>|!:：！]', "", name)

def google_search(query):
    """
    透過 Google Custom Search API 執行搜尋，並返回第一個結果的連結。
    """
    try:
        url = (
            f"https://www.googleapis.com/customsearch/v1"
            f"?key={API_KEY}&cx={SEARCH_ENGINE_ID}&q={query}&gl=tw&hl=zh-TW" # gl=tw 和 hl=zh-TW 鎖定台灣繁體中文結果
        )
        res = requests.get(url)
        if res.status_code == 200:
            items = res.json().get("items")
            # 如果有搜尋結果，返回第一個結果的連結；否則返回 "N"
            return items[0]["link"] if items else "N"
        else:
            # 搜尋 API 返回非 200 狀態碼時的錯誤處理
            print(f"搜尋錯誤: {query} => 狀態碼 {res.status_code}")
            return "N"
    except requests.exceptions.RequestException as e:
        # 處理網路請求相關的異常
        print(f"網路請求失敗: {query} => {e}")
        return "N"
    except Exception as e:
        # 處理其他未知異常
        print(f"搜尋失敗: {query} => {e}")
        return "N"

def load_existing_data(json_file="games.json"):
    """
    載入現有的 games.json 數據，如果檔案不存在則返回空字典。
    """
    if os.path.exists(json_file):
        with open(json_file, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def main():
    """
    主函數：讀取 Excel 數據，更新或新增遊戲資訊，並寫入 JSON 檔案。
    """
    # 載入 Excel 工作簿，data_only=True 確保讀取的是單元格的顯示值而非公式
    try:
        wb = load_workbook("games.xlsx", data_only=True)
        ws = wb.active # 獲取活動工作表
    except FileNotFoundError:
        print("錯誤：找不到 games.xlsx 檔案。請確認檔案是否存在於腳本相同目錄。")
        return

    games_data = load_existing_data() # 載入現有 JSON 數據

    excel_game_names = set()
    # 遍歷 Excel 的每一行，從第二行開始 (假設第一行是標題)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        # 假設遊戲名稱在 Excel 的第一列 (索引 0)
        game_name_from_excel = str(row[0]).strip() if row[0] is not None else ""
        if not game_name_from_excel:
            continue # 如果遊戲名稱為空，跳過此行
        excel_game_names.add(game_name_from_excel)

    # 找出本次 Excel 中新增的遊戲 (不在現有 JSON 數據中的遊戲)
    new_games = excel_game_names - set(games_data.keys())
    print(f"本次 Excel 中共 {len(excel_game_names)} 個遊戲，其中有 {len(new_games)} 個是新增的。")

    # ------------------------------------------------------------------
    # ✅ 階段一：更新所有現有遊戲的禮包碼欄位 (根據 Excel 數據)
    # 這個迴圈會更新 games_data 中已存在的遊戲的禮包碼 URL
    print("\n--- 更新現有遊戲的禮包碼 URL ---")
    for row in ws.iter_rows(min_row=2, values_only=True):
        game_name = str(row[0]).strip() if row[0] is not None else ""
        if not game_name:
            continue
        
        # 假設禮包碼 URL 在 Excel 的第六列 (索引 5)
        excel_gift_url = row[5] if (len(row) > 5 and row[5] is not None) else ""
        
        # 根據 Excel 的值來決定 gift_url
        current_gift_url = excel_gift_url if excel_gift_url else f"gift-codes.html?game={game_name}"
        
        if game_name in games_data:
            # 確保 social 字典存在
            if "social" not in games_data[game_name]:
                games_data[game_name]["social"] = {}
            games_data[game_name]["social"]["禮包碼"] = current_gift_url
            print(f"  - 更新 '{game_name}' 的禮包碼為: {current_gift_url}")
    
    # ------------------------------------------------------------------
    # ✅ 階段二：處理所有遊戲數據 (包括新增的和現有的)
    # 重新遍歷 Excel 數據，處理所有遊戲的詳細資訊
    print("\n--- 處理所有遊戲的詳細資訊 (包括新增與更新) ---")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        game_name = str(row[0]).strip() if row[0] is not None else ""
        if not game_name:
            continue

        # 如果遊戲不在新遊戲列表中，且已經存在於 games_data 中，則跳過 API 搜尋
        # 因為 API 搜尋會比較慢，且假設現有遊戲的 FB/官網/AppStore/巴哈數據是已經處理過的
        # 如果您希望每次都重新搜尋所有遊戲的社交媒體，可以移除這個條件
        if game_name not in new_games and game_name in games_data:
            # 確保 description 欄位也被更新
            # 假設 description 在 Excel 的第八列 (索引 7)
            description_from_excel = row[7] if (len(row) > 7 and row[7] is not None) else ""
            games_data[game_name]['description'] = str(description_from_excel).strip()
            print(f"  - 略過 '{game_name}' 的外部連結搜尋 (已存在)，但更新了簡介。")
            continue # 跳過本次迴圈的其餘部分，處理下一個遊戲


        # 用 clean_filename 處理圖片名稱
        cleaned_game_name = clean_filename(game_name)
        logo = f"images/{cleaned_game_name}.jpg"

        # 假設禮包碼 URL 在 Excel 的第六列 (索引 5)
        excel_gift_url = row[5] if (len(row) > 5 and row[5] is not None) else ""
        gift_url_for_social = excel_gift_url if excel_gift_url else f"gift-codes.html?game={game_name}"

        products = []
        # 從 Excel 的第四列 (索引 3) 開始，每隔兩列讀取商品名稱和價格
        # 這裡假設最多讀取 14 對商品 (3到30)
        for i in range(3, 30, 2):
            if i < len(row) and row[i] is not None:
                pname = str(row[i]).strip()
                if i + 1 < len(row) and row[i + 1] is not None:
                    price = str(row[i + 1]).strip()
                    try:
                        price = int(price)
                        products.append({"name": pname, "price": price})
                    except ValueError:
                        print(f"警告：遊戲 '{game_name}' 商品價格 '{price}' 不是有效數字，已跳過。")
                else:
                    print(f"警告：遊戲 '{game_name}' 商品 '{pname}' 缺少價格，已跳過。")
            else:
                break # 如果商品名稱為空，則停止讀取商品

        # -----------------------------------------------------------
        # 自動搜尋社交媒體連結 (此功能會發送 API 請求，可能會有速率限制)
        print(f"  - 正在搜尋 '{game_name}' 的社交媒體連結...")
        facebook = google_search(f"{game_name} FB")
        website = google_search(f"{game_name} 官方")
        appstore = google_search(f"{game_name} site:apps.apple.com/tw")
        bahamut = google_search(f"{game_name} 巴哈")
        time.sleep(1.2) # 建議每次 API 呼叫之間有延遲，避免觸發速率限制
        # -----------------------------------------------------------

        # 整合所有社交媒體連結
        social_links = {
            "Facebook": facebook,
            "官方網站": website,
            "禮包碼": gift_url_for_social, # 禮包碼使用從 Excel 讀取或生成的 URL
            "App Store": appstore,
            "巴哈姆特": bahamut
        }
        
        # === 新增功能：從 Excel 讀取 description ===
        # 假設 description 在 Excel 的第八列 (索引 7)
        # 請根據您的 Excel 實際欄位位置調整 '7'
        description_from_excel = row[7] if (len(row) > 7 and row[7] is not None) else ""


        # 將所有資訊組合成遊戲數據字典
        games_data[game_name] = {
            "logo": logo,
            "products": products,
            "social": social_links,
            "description": str(description_from_excel).strip() # 將讀取到的簡介存入 JSON
        }

        print(f"✅ '{game_name}' 數據處理完成。")
        time.sleep(0.5) # 處理每個遊戲之間的延遲

    # 寫入更新後的 JSON 檔案
    with open("games.json", "w", encoding="utf-8") as f:
        json.dump(games_data, f, indent=2, ensure_ascii=False)

    print("\n✅ games.json 數據更新完成！")

if __name__ == "__main__":
    main()
